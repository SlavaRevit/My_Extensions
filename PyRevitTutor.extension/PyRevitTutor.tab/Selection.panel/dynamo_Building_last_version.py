import sys
import os
import clr
import sys
import pandas as pd

from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Color
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter
from openpyxl.styles.numbers import FORMAT_PERCENTAGE

localapp = os.getenv(r'LOCALAPPDATA')
sys.path.append(os.path.join(localapp, r'python-3.8.3-embed-amd64\Lib\site-packages'))

clr.AddReference('ProtoGeometry')
from Autodesk.DesignScript.Geometry import *

clr.AddReference("RevitNodes")
import Revit

clr.AddReference("RevitServices")
import RevitServices
from RevitServices.Persistence import DocumentManager
from RevitServices.Transactions import TransactionManager

clr.AddReference("RevitAPI")
clr.AddReference("RevitAPIUI")

import Autodesk
from Autodesk.Revit.DB import *
from Autodesk.Revit.UI import *

doc = DocumentManager.Instance.CurrentDBDocument
uiapp = DocumentManager.Instance.CurrentUIApplication
app = uiapp.Application
uidoc = uiapp.ActiveUIDocument

floors = FilteredElementCollector(doc).OfCategory(BuiltInCategory.OST_Floors). \
    WhereElementIsNotElementType(). \
    ToElements()

walls = FilteredElementCollector(doc).OfCategory(BuiltInCategory.OST_Walls). \
    WhereElementIsNotElementType(). \
    ToElements()

beams = FilteredElementCollector(doc).OfCategory(BuiltInCategory.OST_StructuralFraming). \
    WhereElementIsNotElementType(). \
    ToElements()

levels = FilteredElementCollector(doc).OfCategory(BuiltInCategory.OST_Levels). \
    WhereElementIsNotElementType(). \
    ToElements()
level_floors = {}

for level in levels:
    level_name = level.Name
    level_floors[level_name] = {"מגורים": 0,
                                "מדרגות": 0,
                                "פתחים": 0,
                                "מעליות": 0,
                                "לובי קומתי": 0,
                                "מבואת כניסה": 0,
                                "מועדון דיירים": 0,
                                "שטח צבורי": 0,
                                "מרפסת": 0,
                                "מסתור כביסה": 0,
                                "מרפסת גג": 0,
                                "גג": 0,
                                "גג עליון": 0,
                                "מסחר": 0,
                                "שטח מתחת לקירות": 0,
                                "חדרים טכני ומחסנים": 0,
                                'סה"כ שטח קומה 100%': 0,
                                "פרגולה": 0,
                                "מפלס כפול": 0,
                                "מפולשת": 0
                                }

    for floor in floors:
        floor_id = floor.GetTypeId()
        floor_type = doc.GetElement(floor_id)
        floor_dtm = floor_type.LookupParameter("Duplication Type Mark").AsString()
        floor_TC = floor_type.LookupParameter("Type Comments").AsString()
        floor_level_param = floor.LookupParameter("Level").AsValueString()
        floor_area = floor.LookupParameter("Area").AsDouble() * 0.09290304
        if level_name == floor_level_param:
            if floor_dtm == "Air Stairs":
                level_floors[level_name]['מדרגות'] += floor_area

            elif floor_dtm == "Air Regular":
                level_floors[level_name]["פתחים"] += floor_area

            elif floor_dtm == "Air Elevator":
                level_floors[level_name]["מעליות"] += floor_area

            elif floor_TC == "Lobby":
                level_floors[level_name]["לובי קומתי"] += floor_area

            elif floor_TC == "Main Lobby":
                level_floors[level_name]["מבואת כניסה"] += floor_area

            elif floor_TC == "Club room":
                level_floors[level_name]["מועדון דיירים"] += floor_area

            elif floor_TC in ["Corridor", "Bicycle", "Kindergarten", "Stroller", "Auxiliary"]:
                level_floors[level_name]["שטח צבורי"] += floor_area

            elif floor_TC in ["Balcon", "Balcon Lux"]:
                level_floors[level_name]["מרפסת"] += floor_area

            elif floor_TC == "Mistor Kvisa":
                level_floors[level_name]["מסתור כביסה"] += floor_area

            elif floor_TC in ["Terasa", "Terasa Lux"]:
                level_floors[level_name]["מרפסת גג"] += floor_area

            elif floor_TC == "Roof":
                level_floors[level_name]["גג"] += floor_area

            elif floor_TC == "Trade":
                level_floors[level_name]["מסחר"] += floor_area
            # this need to change to take walls and take it area

            elif floor_TC in ["Warehouse", "Technical", "Garbagee", "Service Cabinet", "Pump"]:
                level_floors[level_name]["חדרים טכני ומחסנים"] += floor_area

            elif floor_dtm == "Total Floor Area":
                level_floors[level_name]['סה"כ שטח קומה 100%'] += floor_area

            elif floor_dtm == "Total Floor Area Pergola":
                level_floors[level_name]["פרגולה"] += floor_area

            elif floor_dtm == "Air Double Level":
                level_floors[level_name]["מפלס כפול"] += floor_area

            elif floor_TC == "overhang":
                level_floors[level_name]["מפולשת"] += floor_area

    for wall in walls:
        wall_id = wall.GetTypeId()
        wall_type = doc.GetElement(wall_id)
        wall_WA = wall_type.LookupParameter("Wall Area").AsValueString()
        wall_level_param = wall.LookupParameter("Base Constraint").AsValueString()
        wall_length = wall.LookupParameter("Length").AsDouble() * 0.3048
        wall_width_param = wall_type.LookupParameter("Width")

        if wall_width_param is not None:
            level_name = level.Name
            wall_width = wall_width_param.AsDouble() * 0.3048
            wall_Area_under = round((wall_length * wall_width), 2)
            if wall_level_param == level_name:
                if wall_WA == "Yes":
                    level_floors[level_name]["שטח מתחת לקירות"] += wall_Area_under

    for beam in beams:
        beam_id = beam.GetTypeId()
        beam_type = doc.GetElement(beam_id)
        beam_level_param = beam.LookupParameter("Reference Level").AsValueString()
        beam_length = beam.LookupParameter("Cut Length").AsDouble() * 0.3048
        beam_dtm = beam_type.LookupParameter("Duplication Type Mark").AsString()
        beam_width_param = beam_type.LookupParameter("b")
        # print(beam_width_param)
        if beam_width_param is not None:
            level_name = level.Name
            beam_width = beam_width_param.AsDouble() * 0.3048
            beam_Area_under = beam_length * beam_width
            if beam_level_param == level_name:
                if beam_dtm == "Balcon":
                    level_floors[level_name]["שטח מתחת לקירות"] += beam_Area_under

row_with_index_arc = pd.Series(IN[2], name="אחוזים למשוקלל")
row_with_index_str = pd.Series(IN[3], name="אחוזים לשלד")

dataframe = pd.DataFrame.from_dict(level_floors,
                                   orient="index",
                                   columns=["מגורים", "מדרגות", "פתחים", "מעליות", "לובי קומתי",
                                            "מבואת כניסה", "מועדון דיירים",
                                            "שטח צבורי",
                                            "מרפסת", "מסתור כביסה", "מרפסת גג", "גג", "גג עליון", "מסחר",
                                            "גג קולונדה", "שטח מתחת לקירות", "חדרים טכני ומחסנים", 'סה"כ שטח קומה 100%',
                                            "פרגולה", "מפלס כפול",
                                            "מפולשת"]).round(1)

df_to_sum = dataframe.loc[:, ["מדרגות", "פתחים", "מעליות", "לובי קומתי",
                              "מבואת כניסה", "מועדון דיירים",
                              "שטח צבורי", "מרפסת", "מסתור כביסה", "מרפסת גג", "גג", "גג עליון", "מסחר",
                              "גג קולונדה", "שטח מתחת לקירות", "חדרים טכני ומחסנים"]].sum(axis=1)

format_str = "{:.0%}"
dataframe.loc['סה"כ 100%'] = dataframe.sum(axis=0)
dataframe = dataframe._append(row_with_index_arc)
dataframe.loc['סה"כ שטח משוקלל'] = dataframe.loc['סה"כ 100%'].mul(row_with_index_arc, axis=0)
dataframe = dataframe._append(row_with_index_str)
dataframe.loc['סה"כ שטח שלד'] = dataframe.loc['סה"כ 100%'].mul(row_with_index_str, axis=0)

dataframe.loc["אחוזים למשוקלל"] = dataframe.loc["אחוזים למשוקלל"].map(format_str.format)
dataframe.loc["אחוזים לשלד"] = dataframe.loc["אחוזים לשלד"].map(format_str.format)

# Compute the sum of the "result" row
result_sum = dataframe.loc['סה"כ שטח משוקלל'].sum()
result_sum_str = dataframe.loc['סה"כ שטח שלד'].sum()

# Assign the sum to the "Total Area" column of the "result" row
dataframe.at['סה"כ שטח משוקלל', 'סה"כ שטח קומה 100%'] = result_sum
dataframe.at['סה"כ שטח שלד', 'סה"כ שטח קומה 100%'] = result_sum_str

dataframe = dataframe.replace(0, "")
dataframe = dataframe.replace("nan%", "")

# export dataframe to excel
writer = pd.ExcelWriter(IN[1], engine='openpyxl')
dataframe.to_excel(writer, sheet_name="שטחים", index=True)

workbook = writer.book
worksheet = writer.sheets['שטחים']
column_index = 1  # column A

sheet = workbook["שטחים"]
min_column = workbook.active.min_column
max_column = workbook.active.max_column
min_row = workbook.active.min_row
max_row = workbook.active.max_row

# setting up formulas
for i in range(min_column + 1, max_column + 1):
    letter = get_column_letter(i)
    sheet[f"{letter}{max_row - 4}"] = f"=SUM({letter}{min_row + 1}:{letter}{max_row - 5})"
    sheet[f"{letter}{max_row - 2}"] = f"={letter}{max_row - 4}*{letter}{max_row - 3}"
    sheet[f"{letter}{max_row}"] = f"={letter}{max_row - 4}*{letter}{max_row - 1}"
    sheet[f"{letter}{max_row - 1}"].number_format = "0%"
    sheet[f"{letter}{max_row - 3}"].number_format = "0%"

col_end_2 = get_column_letter(max_column - 2)
col_total = get_column_letter(max_column - 3)

col_end_4 = get_column_letter(max_column - 4)
col_1 = get_column_letter(min_column + 1)
col_2 = get_column_letter(min_column + 2)

for n in range(min_row + 1, max_row - 4):
    col_letter = get_column_letter(n)
    sheet[f"{col_1}{n}"] = f"={col_total}{n} - SUM({col_2}{n}:{col_end_4}{n})"

sheet[
    f"{col_total}{max_row - 2}"] = f"=SUM({col_1}{max_row - 2}:{col_end_4}{max_row - 2})+SUM({get_column_letter(max_column - 2)}{max_row - 2}:{get_column_letter(max_column)}{(max_row - 2)})"
sheet[
    f"{col_total}{max_row}"] = f"=SUM({col_1}{max_row}:{col_end_4}{max_row})+SUM({get_column_letter(max_column - 2)}{max_row}:{get_column_letter(max_column)}{(max_row)})"

# top row colorized
for col_num, column_title in enumerate(dataframe.columns, 1):
    cell = worksheet.cell(row=1, column=col_num + 1)
    cell.fill = PatternFill(start_color="00CCCCFF", end_color="00CCCCFF", fill_type="solid")

for row_index_ARC in range(1, dataframe.shape[0] + 2):
    cell = worksheet.cell(row=row_index_ARC, column=column_index)
    cell.fill = PatternFill(start_color="00FFFFCC", end_color="00FFFFCC", fill_type="solid")
    cell.alignment = Alignment(horizontal="center", vertical="center")

row_name_ARC = "אחוזים למשוקלל"
row_name_STR = "אחוזים לשלד"
row_name_result = 'סה"כ שטח משוקלל'
row_name_result_str = 'סה"כ שטח שלד'
row_name_100 = 'סה"כ 100%'
row_index_ARC = dataframe.index.get_loc(row_name_ARC) + 2
row_index_STR = dataframe.index.get_loc(row_name_STR) + 2
row_index_result = dataframe.index.get_loc(row_name_result) + 2
row_index_100 = dataframe.index.get_loc(row_name_100) + 2
row_index_result_str = dataframe.index.get_loc(row_name_result_str) + 2

for col_index in range(dataframe.shape[1] + 1):
    col_letter = get_column_letter(col_index + 1)
    cell4 = worksheet.cell(row=row_index_100, column=col_index + 1)
    cell1 = worksheet.cell(row=row_index_ARC, column=col_index + 1)
    cell3 = worksheet.cell(row=row_index_result, column=col_index + 1)
    cell2 = worksheet.cell(row=row_index_STR, column=col_index + 1)
    cell5 = worksheet.cell(row=row_index_result_str, column=col_index + 1)
    cell1.fill = PatternFill(start_color="00CCFFFF", end_color="00CCFFFF", fill_type="solid")
    cell2.fill = PatternFill(start_color="00CCFFFF", end_color="00CCFFFF", fill_type="solid")
    cell3.fill = PatternFill(start_color="00FF8080", end_color="00FF8080", fill_type="solid")
    cell4.fill = PatternFill(start_color="00FFFF00", end_color="00FFFF00", fill_type="solid")  # yellow
    cell5.fill = PatternFill(start_color="00FF8080", end_color="00FF8080", fill_type="solid")

workbook.save(IN[1])

writer.save()

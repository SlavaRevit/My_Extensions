#! python3

import sys
import os
import clr
import sys
import pandas as pd

from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Color
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter
from openpyxl.styles.numbers import FORMAT_PERCENTAGE
from  Autodesk.Revit.UI import TaskDialog, TaskDialogResult


clr.AddReference('ProtoGeometry')
from Autodesk.DesignScript.Geometry import *

clr.AddReference("RevitNodes")
import Revit

clr.AddReference("RevitServices")
import RevitServices
from RevitServices.Persistence import DocumentManager
from RevitServices.Transactions import TransactionManager

clr.AddReference("RevitAPI")
clr.AddReference("RevitAPIUI")

import Autodesk
from Autodesk.Revit.DB import *
from Autodesk.Revit.UI import *

doc = __revit__.ActiveUIDocument.Document
# uiapp = DocumentManager.Instance.CurrentUIApplication
# app = uiapp.Application
# uidoc = uiapp.ActiveUIDocument

floors = FilteredElementCollector(doc).OfCategory(BuiltInCategory.OST_Floors). \
    WhereElementIsNotElementType(). \
    ToElements()

walls = FilteredElementCollector(doc).OfCategory(BuiltInCategory.OST_Walls). \
    WhereElementIsNotElementType(). \
    ToElements()

beams = FilteredElementCollector(doc).OfCategory(BuiltInCategory.OST_StructuralFraming). \
    WhereElementIsNotElementType(). \
    ToElements()

levels = FilteredElementCollector(doc).OfCategory(BuiltInCategory.OST_Levels). \
    WhereElementIsNotElementType(). \
    ToElements()
level_floors = {}

for level in levels:
    level_name = level.Name
    level_floors[level_name] = {
        "חניון": 0,
        "פתחים": 0,
        "מדרגות": 0,
        "רמפות": 0,
        "מעלית": 0,
        "מעבר": 0,
        "לובי": 0,
        "פירים וארונות": 0,
        "חדרים טכני ומחסנים": 0,
        'סה"כ שטח קומה 100%': 0,
        'מפלס כפול': 0
    }

    for floor in floors:
        floor_id = floor.GetTypeId()
        floor_type = doc.GetElement(floor_id)
        floor_dtm = floor_type.LookupParameter("Duplication Type Mark").AsString()
        floor_TC = floor_type.LookupParameter("Type Comments").AsString()
        floor_level_param = floor.LookupParameter("Level").AsValueString()
        floor_area = floor.LookupParameter("Area").AsDouble() * 0.09290304
        if level_name == floor_level_param:
            if floor_dtm == "Air Stairs":
                level_floors[level_name]['מדרגות'] += floor_area

            elif floor_dtm == "Air Regular":
                level_floors[level_name]["פתחים"] += floor_area

            elif floor_dtm == "Rampa":
                level_floors[level_name]["רמפות"] += floor_area

            elif floor_dtm == "Air Elevator":
                level_floors[level_name]["מעלית"] += floor_area

            elif floor_TC == "Lobby":
                level_floors[level_name]["לובי"] += floor_area

            # elif floor_TC in ["Corridor", "Auxiliary"]:
            #     level_floors[level_name]["שטח צבורי"] += floor_area

            elif floor_TC in ["Warehouse", "Technical", "Garbagee", "Service Cabinet", "Pump"]:
                level_floors[level_name]["חדרים טכני ומחסנים"] += floor_area

            elif floor_dtm == "Total Floor Area":
                level_floors[level_name]['סה"כ שטח קומה 100%'] += floor_area

            elif floor_dtm == "Air Double Level":
                level_floors[level_name]["מפלס כפול"] += floor_area


arc_per  = {"חניון":1,
"פתחים":1,
"מדרגות":1,
"רמפות":1,
"מעלית":1,
"מעבר":1,
"לובי":1,
"פירים וארונות":1,
"חדרים טכני ומחסנים":1,
"מפלס כפול":0
};

str_per = {"חניון":1,
"פתחים":1,
"מדרגות":1,
"רמפות":1,
"מעלית":1,
"מעבר":1,
"לובי":1,
"פירים וארונות":1,
"חדרים טכני ומחסנים":1,
"מפלס כפול":0.5
};

translate = {
        "חניון": "Area parking",
        "פתחים": "Elevator",
        "מדרגות": "Opening",
        "רמפות": "Stairs",
        "מעלית": "Ramp",
        "מעבר": "Transition",
        "לובי": "Lobby",
        "פירים וארונות": "Cabinetes",
        "חדרים טכני ומחסנים": "TechRooms",
        'סה"כ שטח קומה 100%': "Total Area",
        'מפלס כפול': "Double level"
    }

row_with_index_arc = pd.Series(arc_per, name="אחוזים למשוקלל")
row_with_index_str = pd.Series(str_per, name="אחוזים לשלד")
row_with_index_translate = pd.Series(translate, name="Translate")

dataframe = pd.DataFrame.from_dict(level_floors,
                                   orient="index",
                                   columns=[
                                       "חניון",
                                       "פתחים",
                                       "מדרגות",
                                       "רמפות",
                                       "מעלית",
                                       "מעבר",
                                       "לובי",
                                       "פירים וארונות",
                                       "חדרים טכני ומחסנים",
                                       'סה"כ שטח קומה 100%',
                                       'מפלס כפול']).round(1)


dataframe = dataframe.append(row_with_index_translate)
format_str = "{:.0%}"
dataframe.loc['סה"כ 100%'] = dataframe.sum(axis=0)
dataframe = dataframe.append(row_with_index_arc)
dataframe.loc['סה"כ שטח משוקלל'] = dataframe.loc['סה"כ 100%'].mul(row_with_index_arc, axis=0)
dataframe = dataframe.append(row_with_index_str)
dataframe.loc['סה"כ שטח שלד'] = dataframe.loc['סה"כ 100%'].mul(row_with_index_str, axis=0)

dataframe.loc["אחוזים למשוקלל"] = dataframe.loc["אחוזים למשוקלל"].map(format_str.format)
dataframe.loc["אחוזים לשלד"] = dataframe.loc["אחוזים לשלד"].map(format_str.format)

# Compute the sum of the "result" row
result_sum = dataframe.loc['סה"כ שטח משוקלל'].sum()
result_sum_str = dataframe.loc['סה"כ שטח שלד'].sum()

# Assign the sum to the "Total Area" column of the "result" row
dataframe.at['סה"כ שטח משוקלל', 'סה"כ שטח קומה 100%'] = result_sum
dataframe.at['סה"כ שטח שלד', 'סה"כ שטח קומה 100%'] = result_sum_str

dataframe = dataframe.replace(0, "")
dataframe = dataframe.replace("nan%", "")



from System.Windows.Forms import Form, Button, Label, TextBox, DialogResult
from System.Drawing import Point

class UserInputDialog(Form):
    def __init__(self):
        self.Text = "User Input"

        # Create labels and text boxes
        self.label = Label()
        self.label.Text = "Please enter your input:"
        self.label.Location = Point(10, 10)
        self.label.AutoSize = True

        self.input_box = TextBox()
        self.input_box.Location = Point(10, 30)
        self.input_box.Width = 200

        # Create OK and Cancel buttons
        self.ok_button = Button()
        self.ok_button.Text = "OK"
        self.ok_button.DialogResult = DialogResult.OK
        self.ok_button.Location = Point(10, 60)

        self.cancel_button = Button()
        self.cancel_button.Text = "Cancel"
        self.cancel_button.DialogResult = DialogResult.Cancel
        self.cancel_button.Location = Point(90, 60)

        # Add controls to the form
        self.Controls.Add(self.label)
        self.Controls.Add(self.input_box)
        self.Controls.Add(self.ok_button)
        self.Controls.Add(self.cancel_button)


# Create an instance of the custom dialog box and display it
dialog = UserInputDialog()
result = dialog.ShowDialog()

if result == DialogResult.OK:
    input_value = dialog.input_box.Text



# export dataframe to excel
writer = pd.ExcelWriter(input_value, engine='openpyxl')
dataframe.to_excel(writer, sheet_name="שטחים", index=True)

workbook = writer.book
worksheet = writer.sheets['שטחים']
column_index = 1  # column A

sheet = workbook["שטחים"]
min_column = workbook.active.min_column
max_column = workbook.active.max_column
min_row = workbook.active.min_row
max_row = workbook.active.max_row

# setting up formulas
for i in range(min_column + 1, max_column + 1):
    letter = get_column_letter(i)
    sheet[f"{letter}{max_row - 4}"] = f"=SUM({letter}{min_row + 1}:{letter}{max_row - 6})"
    sheet[f"{letter}{max_row - 2}"] = f"={letter}{max_row - 4}*{letter}{max_row - 3}"
    sheet[f"{letter}{max_row}"] = f"={letter}{max_row - 4}*{letter}{max_row - 1}"
    sheet[f"{letter}{max_row - 1}"].number_format = "0%"
    sheet[f"{letter}{max_row - 3}"].number_format = "0%"

col_end_2 = get_column_letter(max_column - 2)
col_total = get_column_letter(max_column - 1)

col_end_4 = get_column_letter(max_column - 4)
col_end_2 = get_column_letter(max_column - 2)
col_1 = get_column_letter(min_column + 1)
col_2 = get_column_letter(min_column + 2)

for n in range(min_row + 1, max_row - 5):
    col_letter = get_column_letter(n)
    sheet[f"{col_1}{n}"] = f"={col_total}{n} - SUM({col_2}{n}:{col_end_2}{n})"
    sheet[f"{col_total}{max_row - 2}"] = f"=SUM({col_1}{max_row - 2}:{col_end_2}{max_row - 2})+ ({get_column_letter(max_column)}{max_row-2})"
    sheet[f"{col_total}{max_row}"] = f"=SUM({col_1}{max_row}:{col_end_2}{max_row}) + ({get_column_letter(max_column)}{max_row})"

# top row colorized
for col_num, column_title in enumerate(dataframe.columns, 1):
    cell = worksheet.cell(row=1, column=col_num + 1)
    cell.fill = PatternFill(start_color="00CCCCFF", end_color="00CCCCFF", fill_type="solid")

for row_index_ARC in range(1, dataframe.shape[0] + 2):
    cell = worksheet.cell(row=row_index_ARC, column=column_index)
    cell.fill = PatternFill(start_color="00FFFFCC", end_color="00FFFFCC", fill_type="solid")
    cell.alignment = Alignment(horizontal="center", vertical="center")

row_name_ARC = "אחוזים למשוקלל"
row_name_STR = "אחוזים לשלד"
row_name_result = 'סה"כ שטח משוקלל'
row_name_result_str = 'סה"כ שטח שלד'
row_name_100 = 'סה"כ 100%'
row_index_ARC = dataframe.index.get_loc(row_name_ARC) + 2
row_index_STR = dataframe.index.get_loc(row_name_STR) + 2
row_index_result = dataframe.index.get_loc(row_name_result) + 2
row_index_100 = dataframe.index.get_loc(row_name_100) + 2
row_index_result_str = dataframe.index.get_loc(row_name_result_str) + 2

for col_index in range(dataframe.shape[1] + 1):
    col_letter = get_column_letter(col_index + 1)
    cell4 = worksheet.cell(row=row_index_100, column=col_index + 1)
    cell1 = worksheet.cell(row=row_index_ARC, column=col_index + 1)
    cell3 = worksheet.cell(row=row_index_result, column=col_index + 1)
    cell2 = worksheet.cell(row=row_index_STR, column=col_index + 1)
    cell5 = worksheet.cell(row=row_index_result_str, column=col_index + 1)
    cell1.fill = PatternFill(start_color="00CCFFFF", end_color="00CCFFFF", fill_type="solid")
    cell2.fill = PatternFill(start_color="00CCFFFF", end_color="00CCFFFF", fill_type="solid")
    cell3.fill = PatternFill(start_color="00FF8080", end_color="00FF8080", fill_type="solid")
    cell4.fill = PatternFill(start_color="00FFFF00", end_color="00FFFF00", fill_type="solid")  # yellow
    cell5.fill = PatternFill(start_color="00FF8080", end_color="00FF8080", fill_type="solid")

workbook.save(input_value)

writer.save()
